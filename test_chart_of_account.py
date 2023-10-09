import openpyxl

# Load the Excel file
file_path = './result.xlsx'
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Create a set to store account names
account_names = set()

# Loop through the rows in the Excel sheet to collect account names
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, account_number, is_group, parent_account_name = row
    if name is not None and name != "null":
        account_names.add(name)

# Loop through the rows to check if parent_account_name exists in account_names
for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    name, account_number, is_group, parent_account_name = row
    if parent_account_name is not None and parent_account_name != "null":
        if parent_account_name not in account_names:
            print(f"Row {row_number}: Parent account '{parent_account_name}' does not exist for account '{name}'")
        else:
            print(f"Row {row_number}: [OK]")


# Close the Excel file
workbook.close()
